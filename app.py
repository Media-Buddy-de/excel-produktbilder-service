"""
Excel Produktbilder Service
============================
Web-Service der Excel-Dateien empfängt, Produktbilder herunterlädt 
und einfügt, dann zurückgibt.

Für Make.com optimiert.
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import requests
from io import BytesIO
import base64
import logging
import os

app = Flask(__name__)
CORS(app)

# Logging konfigurieren
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def download_product_image(artikelnummer):
    """
    Versucht Produktbild von 4 verschiedenen URL-Varianten herunterzuladen
    """
    base_url = "https://mb-lightingsolutions.com/wp-content/uploads/2025/11/"
    
    # Die 4 URL-Varianten
    url_variants = [
        f"{base_url}{artikelnummer}_1.jpg",
        f"{base_url}thumbnail_{artikelnummer}_1.jpg",
        f"{base_url}thumbnail_{artikelnummer}_1-scaled.jpg",
        f"{base_url}{artikelnummer}_3-scaled.jpg"
    ]
    
    for url in url_variants:
        try:
            logger.info(f"Versuche: {url}")
            response = requests.get(url, timeout=10)
            
            if response.status_code == 200:
                logger.info(f"✓ Bild gefunden: {url}")
                img = Image.open(BytesIO(response.content))
                return img, url
                
        except Exception as e:
            logger.warning(f"Fehler bei {url}: {str(e)}")
            continue
    
    logger.warning(f"Kein Bild gefunden für: {artikelnummer}")
    return None, None


def compress_image(img, max_width=150, max_height=150, quality=75):
    """Komprimiert Bild für Excel"""
    img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
    
    if img.mode == 'RGBA':
        img = img.convert('RGB')
    
    output = BytesIO()
    img.save(output, format='JPEG', quality=quality, optimize=True)
    output.seek(0)
    
    return Image.open(output)


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({"status": "healthy", "service": "excel-produktbilder"}), 200


@app.route('/process-excel', methods=['POST'])
def process_excel():
    """
    Hauptendpoint: Empfängt Excel, fügt Produktbilder hinzu
    
    Expected JSON:
    {
        "excel_data": "base64_encoded_excel",
        "filename": "original.xlsx"
    }
    
    Returns:
    {
        "success": true,
        "output_file": "base64_encoded_excel",
        "output_filename": "bearbeitet_original.xlsx",
        "stats": {...}
    }
    """
    try:
        data = request.get_json()
        
        if not data or 'excel_data' not in data:
            return jsonify({
                "success": False,
                "error": "Kein Excel-Datei übertragen (excel_data fehlt)"
            }), 400
        
        excel_base64 = data['excel_data']
        original_filename = data.get('filename', 'output.xlsx')
        
        # Base64 dekodieren
        try:
            excel_bytes = base64.b64decode(excel_base64)
        except Exception as e:
            return jsonify({
                "success": False,
                "error": f"Ungültiges Base64-Format: {str(e)}"
            }), 400
        
        # Excel laden
        try:
            wb = openpyxl.load_workbook(BytesIO(excel_bytes))
            ws = wb.active
        except Exception as e:
            return jsonify({
                "success": False,
                "error": f"Ungültige Excel-Datei: {str(e)}"
            }), 400
        
        # Spalte "Artikel Nr." finden (flexibel)
        header_row = ws[1]
        artikel_col = None
        
        for cell in header_row:
            if cell.value:
                cell_text = str(cell.value).strip().lower()
                # Suche nach verschiedenen Varianten
                if "artikel" in cell_text and ("nr" in cell_text or "nummer" in cell_text):
                    artikel_col = cell.column
                    logger.info(f"Artikel-Spalte gefunden: '{cell.value}' in Spalte {cell.column}")
                    break
        
        if not artikel_col:
            return jsonify({
                "success": False,
                "error": "Spalte 'Artikel Nr.' nicht gefunden. Verfügbare Spalten: " + ", ".join([str(c.value) for c in header_row if c.value])
            }), 400
        
        # Neue Spalte A einfügen für Bilder - ALLE Spalten rutschen nach rechts!
        ws.insert_cols(1)  # Fügt neue Spalte A ein
        ws.cell(row=1, column=1, value="Artikelbild")  # Header
        
        # Artikel-Spalte ist jetzt um 1 nach rechts gerutscht!
        artikel_col += 1
        
        bild_col = 1  # Bilder kommen in Spalte A
        
        # Statistiken
        stats = {
            "total_rows": ws.max_row - 1,
            "successful_images": 0,
            "failed_images": 0,
            "errors": []
        }
        
        # Durch alle Zeilen gehen
        for row_idx in range(2, ws.max_row + 1):
            artikelnummer = ws.cell(row=row_idx, column=artikel_col).value
            
            if not artikelnummer:
                continue
            
            # Bild herunterladen
            img, img_url = download_product_image(str(artikelnummer))
            
            if img:
                try:
                    # Komprimieren
                    img_compressed = compress_image(img)
                    
                    # Als BytesIO speichern
                    img_bytes = BytesIO()
                    img_compressed.save(img_bytes, format='JPEG')
                    img_bytes.seek(0)
                    
                    # In Excel einfügen
                    xl_image = XLImage(img_bytes)
                    xl_image.width = 100
                    xl_image.height = 100
                    
                    cell_address = ws.cell(row=row_idx, column=bild_col).coordinate
                    ws.add_image(xl_image, cell_address)
                    
                    # Zeilenhöhe anpassen
                    ws.row_dimensions[row_idx].height = 75
                    
                    stats["successful_images"] += 1
                    logger.info(f"✓ Zeile {row_idx}: Bild eingefügt ({artikelnummer})")
                    
                except Exception as e:
                    stats["failed_images"] += 1
                    stats["errors"].append(f"Zeile {row_idx}: Fehler beim Einfügen - {str(e)}")
                    logger.error(f"Fehler bei Zeile {row_idx}: {str(e)}")
            else:
                stats["failed_images"] += 1
                stats["errors"].append(f"Zeile {row_idx}: Kein Bild gefunden für {artikelnummer}")
        
        # Spaltenbreite anpassen
        ws.column_dimensions[openpyxl.utils.get_column_letter(bild_col)].width = 15
        
        # Excel speichern
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Sende Excel direkt als Binary-Download (nicht als JSON!)
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"bearbeitet_{original_filename}"
        )
        
    except Exception as e:
        logger.error(f"Unerwarteter Fehler: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"Server-Fehler: {str(e)}"
        }), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
